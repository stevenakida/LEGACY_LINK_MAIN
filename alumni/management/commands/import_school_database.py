from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify

from alumni.models import School

# Each data sheet in the master spreadsheet shares the same 15-column
# layout: legacy_school_id, official_school_name, education_level, region,
# district, ward, ownership, source_code, source_id, normalized_key,
# alias_examples, status, validation_status, source, source_url.
SHEET_TO_SCHOOL_TYPE = [
    ('Primary_Schools', 'primary'),
    ('O_Level_Secondary', 'secondary'),
    ('A_Level_Schools', 'high_school'),
    ('Universities_Colleges', 'university'),
]

DEFAULT_PATH = settings.BASE_DIR.parent / 'LegacyLink_Africa_Tanzania_Education_Master_Database.xlsx'
CHUNK_SIZE = 2000


class Command(BaseCommand):
    help = (
        'Import/refresh the primary, secondary (O-Level), high school (A-Level) '
        'and university records from the LegacyLink Tanzania education master '
        'spreadsheet. Safe to re-run: rows are matched by their spreadsheet '
        'legacy_school_id, so re-running updates existing records instead of '
        'duplicating them.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            'path', nargs='?', default=str(DEFAULT_PATH),
            help='Path to the .xlsx master database (defaults to the copy in the project root).'
        )
        parser.add_argument(
            '--purge-legacy-seed', action='store_true',
            help='Delete School rows with no external_id (the old hand-typed seed_schools list) after import.'
        )

    def handle(self, *args, **options):
        path = Path(options['path'])
        if not path.exists():
            raise CommandError(f'Spreadsheet not found at {path}')

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        total_created = 0
        total_updated = 0
        total_skipped = 0

        for sheet_name, school_type in SHEET_TO_SCHOOL_TYPE:
            if sheet_name not in wb.sheetnames:
                raise CommandError(f'Expected sheet "{sheet_name}" not found in {path.name}')

            rows = []
            for row in wb[sheet_name].iter_rows(min_row=3, values_only=True):
                if not row or not row[0] or not row[1]:
                    continue
                legacy_id, name = row[0], row[1]
                region = (row[3] or '').strip()
                district = (row[4] or '').strip()
                rows.append({
                    'external_id': str(legacy_id).strip(),
                    'name': str(name).strip(),
                    'region': region,
                    'district': district,
                    'school_type': school_type,
                })

            created, updated, skipped = self._upsert(rows)
            total_created += created
            total_updated += updated
            total_skipped += skipped
            self.stdout.write(f'{sheet_name}: {created} created, {updated} updated, {skipped} unchanged')

        self.stdout.write(self.style.SUCCESS(
            f'Done. {total_created} created, {total_updated} updated, {total_skipped} unchanged.'
        ))

        if options['purge_legacy_seed']:
            legacy_qs = School.objects.filter(external_id__isnull=True)
            count = legacy_qs.count()
            legacy_qs.delete()
            self.stdout.write(self.style.WARNING(
                f'Purged {count} legacy seed_schools rows (no external_id). '
                'Any user profile pointing at one of those now shows no school selected.'
            ))

    def _upsert(self, rows):
        created = updated = skipped = 0
        for i in range(0, len(rows), CHUNK_SIZE):
            chunk = rows[i:i + CHUNK_SIZE]
            existing = {
                s.external_id: s
                for s in School.objects.filter(external_id__in=[r['external_id'] for r in chunk])
            }

            to_create = []
            to_update = []
            for r in chunk:
                obj = existing.get(r['external_id'])
                if obj is None:
                    slug_source = f"{r['name']}-{r['school_type']}-{r['external_id']}"
                    to_create.append(School(
                        slug=slugify(slug_source)[:300],
                        **r,
                    ))
                    continue

                changed = False
                for field in ('name', 'region', 'district', 'school_type'):
                    if getattr(obj, field) != r[field]:
                        setattr(obj, field, r[field])
                        changed = True
                if changed:
                    to_update.append(obj)
                else:
                    skipped += 1

            if to_create:
                School.objects.bulk_create(to_create, ignore_conflicts=True)
                created += len(to_create)
            if to_update:
                School.objects.bulk_update(to_update, ['name', 'region', 'district', 'school_type'])
                updated += len(to_update)

        return created, updated, skipped
